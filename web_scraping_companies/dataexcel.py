from openpyxl import Workbook
# Import the companies class
from companies_class.amarilo_company import AmariloProject
from companies_class.bolivar_company import BolivarProjects
from companies_class.prodesa_company import ProdesaProject
from companies_class.colpatria_company import ColpatriaProject
from companies_class.conaltura_company import ConalturaProject
from companies_class.marval_company import MarvalProject
from companies_class.others_company import OtherProjects
from companies_class.arenas_inmobiliarias_company import ArenasProjects
from track_logs.logs import track_logs



# ALL PROJECTS CLASS
amarilo_projects = AmariloProject()
marval_projects = MarvalProject()
bolivar_projects = BolivarProjects()
prodesa_projects = ProdesaProject()
colpatria_projects = ColpatriaProject()
conaltura_projects = ConalturaProject()
arenas_inmobiliarias_projects = ArenasProjects()
others_projects = OtherProjects()

# open Excel file
#workbook = openpyxl.load_workbook("projects.xlsx")
# Create a new workbook and a new sheet
workbook = Workbook()
ws = workbook.active

# Title first sheet
ws.title = "amarilo"

# Create sheets with the name of companies
# worksheet = by default is 'amarilo'
worksheet1 = workbook.create_sheet("marval")
worksheet2 = workbook.create_sheet("arenas_inmobiliaria")
worksheet3 = workbook.create_sheet("bolivar")
worksheet4 = workbook.create_sheet("colpatria")
worksheet5 = workbook.create_sheet("conaltura")
worksheet6 = workbook.create_sheet("prodesa")
worksheet7 = workbook.create_sheet("others")

# tuple of first row
first_row = (
            "name", "logo",	"location",	"city", "company",	"address",	
            "contact", "area", "price", "type",	"img_url",	"description",	"url_website"
)
# WRITE THE SCRAPED DATA FOR EACH SHEET
# Amarilo
worksheet = workbook["amarilo"]
worksheet.append(first_row)
row_cell = 2
for record_project in amarilo_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
            
        worksheet.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
track_logs(f"{worksheet} finish")

# Marval
unique_marval_name_project = []
worksheet1 = workbook["marval"]
worksheet1.append(first_row)
row_cell = 2
for record_project in marval_projects.get_projects_by_soledad():
    colum_cell = 1
    unique_marval_name_project.append(record_project["name"])
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet1.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
for record_project in marval_projects.get_projects_by_barranquilla():
    colum_cell = 1
    if record_project["name"] not in unique_marval_name_project:
        unique_marval_name_project.append(record_project["name"])
        for data in record_project:
            try:
                value_field = record_project[data].strip()
            except AttributeError:
                value_field = record_project[data]
            worksheet1.cell(row=row_cell, column=colum_cell, value=value_field)
            colum_cell += 1
        row_cell += 1
track_logs(f"{worksheet1} finish")

# Arenas Inmobiliaria
worksheet2 = workbook["arenas_inmobiliaria"]
worksheet2.append(first_row)
row_cell = 2
for record_project in arenas_inmobiliarias_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet2.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
track_logs(f"{worksheet2} finish")

# Bolivar
worksheet3 = workbook["bolivar"]
worksheet3.append(first_row)
row_cell = 2
for record_project in bolivar_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet3.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
track_logs(f"{worksheet3} finish")

# Colpatria
worksheet4 = workbook["colpatria"]
worksheet4.append(first_row)
row_cell = 2
for record_project in colpatria_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet4.cell(row=row_cell, column=colum_cell, value=value_field )
        colum_cell += 1
    row_cell += 1
track_logs(f"{worksheet4} finish")

# Conaltura
worksheet5 = workbook["conaltura"]
worksheet5.append(first_row)
row_cell = 2
for record_project in conaltura_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet5.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
track_logs(f"{worksheet5} finish")

# Prodesa
worksheet6 = workbook["prodesa"]
worksheet6.append(first_row)
row_cell = 2
for record_project in prodesa_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet6.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1

track_logs(f"{worksheet6} finish")

# Others
worksheet7 = workbook["others"]
worksheet7.append(first_row)
row_cell = 2
for record_project in others_projects.ACF:
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet7.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
for record_project in others_projects.APIROS:
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet7.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
for record_project in others_projects.COCONCRETO:
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet7.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
for record_project in others_projects.OPINA_CIA:
    colum_cell = 1
    for data in record_project:
        try:
            value_field = record_project[data].strip()
        except AttributeError:
            value_field = record_project[data]
        worksheet7.cell(row=row_cell, column=colum_cell, value=value_field)
        colum_cell += 1
    row_cell += 1
track_logs(f"{worksheet7} finish")

# Save the changes
workbook.save("saved_documents/projects.xlsx")








