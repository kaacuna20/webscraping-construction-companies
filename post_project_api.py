import requests
import os
import re
from natsort import natsorted
from dotenv import load_dotenv
import openpyxl

load_dotenv(".env")

logos_path = rf"{os.getenv('LOGOS_PATH ')}\static\images\logos1"
background_path = rf"{os.getenv('BG_PATH')}\static\images\background1"

dir_list_bg = os.listdir(background_path)
dir_list_logo = os.listdir(logos_path)

# Create a copy from projects.xlsx
workbook = openpyxl.load_workbook("projects_override.xlsx")

for company_sheet in workbook.sheetnames:
    # open the sheet to override the columns of url_img(column=12) and logo(column=2) row by row
    worksheet = workbook[company_sheet]
    # Using regular expression, create a list with the files with the company_sheet filtered from the list
    # that os.listdir() got from my directories  /logos1 and /background1
    list_bg_per_company = [bg for bg in dir_list_bg if re.compile(rf"{company_sheet}", re.IGNORECASE).search(bg)]
    list_logo_per_company = [logo for logo in dir_list_logo if re.compile(rf"{company_sheet}", re.IGNORECASE).search(logo)]
    # sort that list in ascending form
    sorted_bg_per_company = natsorted(list_bg_per_company)
    sorted_logo_per_company = natsorted(list_logo_per_company)

    # Override the column img_url adding the path /background1
    for row in range(len(sorted_bg_per_company)):
        # Here replace '\' with '/' because my OS is windows and the path recognize with '\'
        # But when I update the images, in my folder of static recognize the path with '/'
        path_bg = rf"static\images\background1\{sorted_bg_per_company[row]}".replace('\\', "/")
        worksheet.cell(row=row + 2, column=12, value=path_bg)

    # Override the column logo adding the path /logos1
    for row in range(len(sorted_logo_per_company)):
        path_logo = rf"static\images\logos1\{sorted_logo_per_company[row]}".replace('\\', "/")
        worksheet.cell(row=row + 2, column=2, value=path_logo)


workbook.save("projects_override.xlsx")
# SEND POST REQUEST USING THE API THAT WAS CREATED IN PROJECT "HOUSE PROJECT SEARCHING"
# create a list of tag column
list_item = ["name", "logo", "location", "city", "company", "address", "url_map", "contact",
             "area", "price", "type", "img_url", "description", "url_website"]

# get a valid apikey and the endpoint to make the 'POST' request
API_KEY = os.getenv("ADMI_APIKEY")

# This is the server, this can change
SERVER_LOCALHOST = os.getenv("LOCALHOST")
URL = f"{SERVER_LOCALHOST}/api/add"

# create a dictionary of parameters according to API Documentation
parameters = {}

for company_sheet in workbook.sheetnames:
    # Select the sheets from the list of workbook.sheetnames
    # List of sheetnames: ['amarilo', 'marval', 'arenas_inmobiliaria', 'bolivar', 'colpatria', 'conaltura', 'prodesa', 'others']
    worksheet = workbook[company_sheet]
    # create a loop for since the first until last row of each sheet
    for row_record_project in range(2, worksheet.max_row + 1):
        # create a loop for since the first until last column of each record
        for item in range(0, worksheet.max_column):
            # write the parameters with the values getting from projects_override.xlsx
            parameters["api_key"] = API_KEY
            parameters[list_item[item]] = worksheet.cell(row=row_record_project, column=item + 1).value
        # Sent the POST requests
        response = requests.post(url=URL, params=parameters)
        # watch the response
        print(response.text)
