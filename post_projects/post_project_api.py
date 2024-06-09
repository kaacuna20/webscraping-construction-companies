import requests
import os
from dotenv import load_dotenv
import openpyxl
from track_logs.logs import track_logs

load_dotenv(".env")


# Create a copy from projects.xlsx
workbook = openpyxl.load_workbook("overwrite_projects.xlsx")

# SEND POST REQUEST USING THE API THAT WAS CREATED IN PROJECT "HOUSE PROJECT SEARCHING"
# create a list of tag column
list_item = ["name", "logo", "location", "city", "company", "address", "url_map", "contact",
             "area", "price", "type", "img_url", "description", "url_website"]

# get a valid apikey and the endpoint to make the 'POST' request
API_KEY = os.getenv("PUBLIC_API_KEY")
headers = {
    "Api-key": API_KEY
}

# This is the server, this can change
URL = "http://localhost/api/v1/add-project"

# create a dictionary of parameters according to API Documentation
parameters = {}

for company_sheet in workbook.sheetnames:
    # Select the sheets from the list of workbook.sheetnames
    # List of sheetnames: ['amarilo', 'marval', 'arenas_inmobiliaria', 'bolivar', 'colpatria', 'conaltura', 'prodesa', 'others']
    worksheet = workbook[company_sheet]
    track_logs(worksheet)
    # create a loop for since the first until last row of each sheet
    for row_record_project in range(2, worksheet.max_row + 1):
        # create a loop for since the first until last column of each record
        for item in range(0, worksheet.max_column):
            # write the parameters with the values getting from projects_override.xlsx
            parameters[list_item[item]] = worksheet.cell(row=row_record_project, column=item + 1).value
        try:    
            # Sent the POST requests
            response = requests.post(url=URL, params=parameters, headers=headers)
            # watch the response
            track_logs(response.text)
        except Exception:
            track_logs("Server no avaible!")
