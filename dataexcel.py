import openpyxl
# Import the companies class
from companies_class.amarilo_company import AmariloProject
from companies_class.bolivar_company import BolivarProjects
from companies_class.prodesa_company import ProdesaProject
from companies_class.colpatria_company import ColpatriaProject
from companies_class.conaltura_company import ConalturaProject
from companies_class.marval_company import MarvalProject
from companies_class.others_company import OtherProjects
from companies_class.arenas_inmobiliarias_company import ArenasProjects


# ALL PROJECTS CLASS
amarilo_projects = AmariloProject()
marval_projects = MarvalProject()
bolivar_projects = BolivarProjects()
prodesa_projects = ProdesaProject()
colpatria_projects = ColpatriaProject()
conaltura_projects = ConalturaProject()
arenas_inmobiliarias_projects = ArenasProjects()
others_projects = OtherProjects()

# open Excel file
workbook = openpyxl.load_workbook("projects.xlsx")

# Create sheets with the name of companies
# worksheet = by default is 'amarilo'
worksheet1 = workbook.create_sheet("marval")
worksheet2 = workbook.create_sheet("arenas_inmobiliaria")
worksheet3 = workbook.create_sheet("bolivar")
worksheet4 = workbook.create_sheet("colpatria")
worksheet5 = workbook.create_sheet("conaltura")
worksheet6 = workbook.create_sheet("prodesa")
worksheet7 = workbook.create_sheet("others")


# WRITE THE DATA FOR EACH SHEET
# Amarilo
worksheet = workbook["amarilo"]
row_cell = 2
for record_project in amarilo_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        worksheet.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Conaltura
worksheet5 = workbook["conaltura"]
row_cell = 2
for record_project in conaltura_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        worksheet5.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Bolivar
worksheet3 = workbook["bolivar"]
row_cell = 2
for record_project in bolivar_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        worksheet3.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Arenas Inmobiliaria
worksheet2 = workbook["arenas_inmobiliaria"]
row_cell = 2
for record_project in arenas_inmobiliarias_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        worksheet2.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Colpatria
worksheet4 = workbook["colpatria"]
row_cell = 2
for record_project in colpatria_projects.get_projects():
    colum_cell = 1
    for data in record_project:
        worksheet4.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Prodesa
worksheet6 = workbook["prodesa"]
row_cell = 2
for record_project in prodesa_projects.get_project_by_soledad():
    colum_cell = 1
    for data in record_project:
        worksheet6.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1
for record_project in prodesa_projects.get_project_by_barranquilla():
    colum_cell = 1
    for data in record_project:
        worksheet6.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Marval
worksheet1 = workbook["marval"]
row_cell = 2
for record_project in marval_projects.get_projects_by_soledad():
    colum_cell = 1
    for data in record_project:
        worksheet1.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1
for record_project in marval_projects.get_projects_by_barranquilla():
    colum_cell = 1
    for data in record_project:
        worksheet1.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Others
worksheet7 = workbook["others"]
row_cell = 2
for record_project in others_projects.ACF:
    colum_cell = 1
    for data in record_project:
        worksheet7.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1
for record_project in others_projects.APIROS:
    colum_cell = 1
    for data in record_project:
        worksheet7.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1
for record_project in others_projects.COCONCRETO:
    colum_cell = 1
    for data in record_project:
        worksheet7.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1
for record_project in others_projects.OPINA_CIA:
    colum_cell = 1
    for data in record_project:
        worksheet7.cell(row=row_cell, column=colum_cell, value=record_project[data].strip())
        colum_cell += 1
    row_cell += 1

# Save the changes
workbook.save("projects.xlsx")







