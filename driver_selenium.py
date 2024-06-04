import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, InvalidArgumentException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from track_logs.logs import track_logs
import os
import time


def donwload_images_from_urls(excel_file:str):
    # CREATE THE BOT WITH SELENIUM TO WRITE THE DATA SCRAPED IN THE FORM GOOGLE
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')

    # Setup ChromeDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    workbook = openpyxl.load_workbook(excel_file)
    # list of name of each sheet of Excel file
    track_logs(workbook.sheetnames)
    # List of sheetnames: ['amarilo', 'marval', 'arenas_inmobiliaria', 'bolivar', 'colpatria', 'conaltura', 'prodesa', 'others']

    # main path to save images
    path = "saved_data/static/images/img-projects" 
    
    iter_company = 1
    for sheet_project in workbook.sheetnames:
        
        # Select the sheets from the list of workbook.sheetnames
        worksheet = workbook[sheet_project]
        track_logs(worksheet)
        
        try:
            # create the folder acording the name of company
            os.makedirs(f"{path}/background/{sheet_project}")
            track_logs(f"directory {path}/background/{sheet_project} created")
        except OSError:
            pass
        
        # get the values of column 1 (item=name) and column 12 (item=img_url) and add them in a dictionary
        dict_img_url = {worksheet.cell(row=row_cell, column=1).value: worksheet.cell(row=row_cell, column=12).value for row_cell in range(2,  worksheet.max_row + 1)}
        # get the values of column 1 (item=name) and column 2 (item=logo) and add them in a dictionary
        dict_logo = {worksheet.cell(row=row_cell, column=1).value: worksheet.cell(row=row_cell, column=2).value for row_cell in range(2,  worksheet.max_row + 1)}
        # now we iterate on both dictionaries, to get the url's, open driver and download the html element tag_name='img'
        iter_img = 1
        for name_project in dict_img_url:
            
            try:
                driver.get(url=dict_img_url[name_project])
                time.sleep(2)
                img = driver.find_element(By.TAG_NAME, value="img")

                with open(f'{path}/background/{sheet_project}/{name_project.strip()}.png', 'wb') as file:
                    file.write(img.screenshot_as_png)

            except InvalidArgumentException:
                track_logs(f"bg {sheet_project}-{name_project} could not download")

            except NoSuchElementException:
                track_logs(f"bg {sheet_project}-{name_project}  could not download")

            except WebDriverException:
                track_logs(f"bg {sheet_project}-{name_project}  no found")

            iter_img += 1

        iter_logo = 1
        for name_project in dict_logo:
            try:
                os.makedirs(f"{path}/logos/{sheet_project}")
                track_logs(f"directory {path}/logos/{sheet_project} created")
            except OSError:
                pass
            try:
                driver.get(url=dict_logo[name_project])
                time.sleep(2)
                img = driver.find_element(By.TAG_NAME, value="img")

                with open(f'{path}/logos/{sheet_project}/{name_project.strip()}.png', 'wb') as file:
                    file.write(img.screenshot_as_png)

            except InvalidArgumentException:
                track_logs(f"logo {sheet_project}-{name_project} could not download")

            except NoSuchElementException:
                track_logs(f"logo {sheet_project}-{name_project}  could not download")

            except WebDriverException:
                track_logs(f"logo {sheet_project}-{name_project} no found")

            iter_logo += 1
        iter_company += 1

    driver.close()










