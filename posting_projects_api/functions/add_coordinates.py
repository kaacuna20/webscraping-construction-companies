import openpyxl
from .google_maps_api import GoogleMap

def add_coordinates_to_excelfile(excelfile: str, list_items: list) -> list:            
    workbook = openpyxl.load_workbook(excelfile) 
    list_projects = []

    for sheet_project in workbook.sheetnames:
        
        worksheet = workbook[sheet_project]
               
        for row_record_project in range(2, worksheet.max_row + 1):
            project ={}
            for item in range(len(list_items)):
                # write the parameters with the values getting from projects_override.xlsx
                project[list_items[item]] = worksheet.cell(row=row_record_project, column=item + 1).value
            
            text_coordinates = f"{project['name']}, {project['location']}, {project['city']}, Atlántico"
            project["latitude"] = GoogleMap().get_coordinates(text_coordinates)[0]
            project["longitude"] = GoogleMap().get_coordinates(text_coordinates)[1]
            list_projects.append(project)
            
    return list_projects